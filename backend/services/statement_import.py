import csv
import datetime
import io

from dateutil import parser as date_parser
from openpyxl import load_workbook

from models.transaction import EXPENSE, INCOME

HEADER_ALIASES = {
    "date": {"date", "txn date", "transaction date", "value date"},
    "description": {
        "description", "details", "transaction details", "particulars",
        "narration", "remark", "remarks", "comment",
    },
    "amount": {"amount", "amount (inr)", "amt", "amount(rs)", "amount (rs)"},
    "type": {"type", "transaction type", "dr/cr", "debit/credit", "cr/dr"},
    "debit": {"debit", "debit amount", "withdrawal", "withdrawal amt", "paid"},
    "credit": {"credit", "credit amount", "deposit", "deposit amt", "received"},
}

DEBIT_WORDS = {"debit", "dr", "paid", "withdraw", "withdrawal", "spent", "sent"}
CREDIT_WORDS = {"credit", "cr", "received", "deposit", "refund"}

CATEGORY_KEYWORDS = {
    "Food": ["swiggy", "zomato", "food", "restaurant", "cafe", "domino", "pizza",
              "starbucks", "eatery", "dine", "bakery"],
    "Transport": ["uber", "ola", "irctc", "metro", "fuel", "petrol", "diesel",
                   "rapido", "fastag", "parking", "railway"],
    "Shopping": ["amazon", "flipkart", "myntra", "ajio", "meesho", "mall", "store", "mart"],
    "Bills": ["electricity", "recharge", "broadband", "dth", "water bill", "gas bill",
               "mobile bill", "postpaid", "prepaid", "wifi", "bill payment"],
    "Entertainment": ["netflix", "prime video", "hotstar", "spotify", "bookmyshow",
                        "movie", "pvr", "inox"],
    "Health": ["pharmacy", "hospital", "apollo", "medplus", "clinic", "diagnostic", "medical"],
}


def categorize(description):
    text = (description or "").lower()
    for category, keywords in CATEGORY_KEYWORDS.items():
        if any(keyword in text for keyword in keywords):
            return category
    return "Other"


def _normalize_header(name):
    return " ".join(str(name or "").strip().lower().split())


def _map_headers(headers):
    normalized = [_normalize_header(h) for h in headers]
    column_index = {}
    for field, aliases in HEADER_ALIASES.items():
        for i, header in enumerate(normalized):
            if header in aliases:
                column_index[field] = i
                break
    return column_index


def _parse_amount(raw):
    if raw is None:
        return None
    text = str(raw).strip().replace(",", "").replace("₹", "").replace("Rs.", "").replace("INR", "")
    if not text:
        return None
    negative = text.startswith("-") or (text.startswith("(") and text.endswith(")"))
    text = text.strip("-()").strip()
    try:
        value = float(text)
    except ValueError:
        return None
    return -value if negative else value


def _parse_date(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime.datetime):
        return raw.date()
    if isinstance(raw, datetime.date):
        return raw
    try:
        return date_parser.parse(str(raw), dayfirst=True).date()
    except (ValueError, TypeError, date_parser.ParserError):
        return None


def _rows_to_transactions(headers, rows):
    columns = _map_headers(headers)
    if "date" not in columns:
        return [], ["Couldn't find a date column in this file."]
    has_amount_type = "amount" in columns
    has_debit_credit = "debit" in columns or "credit" in columns
    if not has_amount_type and not has_debit_credit:
        return [], ["Couldn't find an amount column in this file."]

    transactions = []
    skipped = 0

    for row in rows:
        if not any(cell not in (None, "") for cell in row):
            continue

        date_value = _parse_date(row[columns["date"]] if columns["date"] < len(row) else None)
        description = ""
        if "description" in columns and columns["description"] < len(row):
            description = str(row[columns["description"]] or "").strip()

        amount = None
        type_ = None

        if has_debit_credit:
            debit_val = _parse_amount(row[columns["debit"]]) if "debit" in columns and columns["debit"] < len(row) else None
            credit_val = _parse_amount(row[columns["credit"]]) if "credit" in columns and columns["credit"] < len(row) else None
            if debit_val:
                amount, type_ = abs(debit_val), EXPENSE
            elif credit_val:
                amount, type_ = abs(credit_val), INCOME
        else:
            raw_amount = _parse_amount(row[columns["amount"]]) if columns["amount"] < len(row) else None
            if raw_amount is not None:
                type_word = ""
                if "type" in columns and columns["type"] < len(row):
                    type_word = str(row[columns["type"]] or "").strip().lower()
                if type_word in DEBIT_WORDS:
                    amount, type_ = abs(raw_amount), EXPENSE
                elif type_word in CREDIT_WORDS:
                    amount, type_ = abs(raw_amount), INCOME
                elif raw_amount < 0:
                    amount, type_ = abs(raw_amount), EXPENSE
                elif raw_amount > 0:
                    amount, type_ = raw_amount, INCOME

        if date_value is None or amount is None or not amount:
            skipped += 1
            continue

        transactions.append({
            "date": date_value,
            "description": description,
            "amount": amount,
            "type": type_,
        })

    errors = [f"Skipped {skipped} row(s) that couldn't be read."] if skipped else []
    return transactions, errors


def parse_statement(file_storage):
    filename = (file_storage.filename or "").lower()
    raw = file_storage.read()

    if filename.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return [], ["The file is empty."]
        return _rows_to_transactions(rows[0], rows[1:])

    if filename.endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], ["The file is empty."]
        return _rows_to_transactions(list(rows[0]), [list(r) for r in rows[1:]])

    return [], ["Unsupported file type — upload a .csv or .xlsx statement export."]
